from flask import Flask, render_template, request, redirect, url_for, session, send_file, jsonify, flash
import os
import numpy as np
from PIL import Image
import tensorflow as tf
from tensorflow.keras.applications.mobilenet_v2 import preprocess_input, decode_predictions
import openpyxl
from diet_data import generate_diet_plan, get_size_category, get_exact_size
import datetime

# Helper function to convert 24-hour time to 12-hour format
def convert_to_12_hour_format(time_24):
    """Convert 24-hour format time (HH:MM) to 12-hour format with AM/PM"""
    try:
        if not time_24 or ':' not in time_24:
            return time_24
        
        hour, minute = time_24.split(':')
        hour = int(hour)
        
        if hour == 0:
            return f"12:{minute} AM"
        elif hour < 12:
            return f"{hour}:{minute} AM"
        elif hour == 12:
            return f"12:{minute} PM"
        else:
            return f"{hour - 12}:{minute} PM"
    except:
        return time_24

app = Flask(__name__)
app.secret_key = "secretkey"

# ======================================================
# CONFIG
# ======================================================

UPLOAD_FOLDER = 'static/uploads'
EXCEL_FILE = 'users.xlsx'

app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER

ADMIN_EMAIL = "admin@gmail.com"
ADMIN_PASSWORD = "admin123"

if not os.path.exists(UPLOAD_FOLDER):
    os.makedirs(UPLOAD_FOLDER)

# ======================================================
# CREATE EXCEL FILE
# ======================================================

if not os.path.exists(EXCEL_FILE):
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = "Users"
    sheet.append(["Username", "Email", "Password"])
    wb.save(EXCEL_FILE)

# ======================================================
# LOAD MODEL
# ======================================================

model = tf.keras.applications.MobileNetV2(weights='imagenet')

# ======================================================
# USER FUNCTIONS
# ======================================================

def register_user(username, email, password):
    try:
        wb = openpyxl.load_workbook(EXCEL_FILE)
        sheet = wb.active
        
        email_to_check = email.strip().lower()

        for row in sheet.iter_rows(min_row=2, values_only=True):
            if not row or len(row) < 2:
                continue
            db_email = str(row[1]).strip().lower() if row[1] is not None else ""
            if db_email == email_to_check:
                return False

        sheet.append([
            username.strip(),
            email_to_check,
            password
        ])

        wb.save(EXCEL_FILE)
        return True
    except Exception as e:
        print(f"Error in register_user: {e}")
        return False


def check_user(email, password):
    try:
        wb = openpyxl.load_workbook(EXCEL_FILE)
        sheet = wb.active

        for row in sheet.iter_rows(min_row=2, values_only=True):
            if not row or len(row) < 3:
                continue
            
            db_user = str(row[0]).strip() if row[0] is not None else ""
            db_email = str(row[1]).strip().lower() if row[1] is not None else ""
            db_pass = str(row[2]).strip() if row[2] is not None else ""

            if db_email == email.strip().lower() and db_pass == password.strip():
                return db_user if db_user else "User"

        return None
    except Exception as e:
        print(f"Error in check_user: {e}")
        return None

# ======================================================
# PREDICT BREED
# ======================================================

def predict_breed(img_path):

    img = Image.open(img_path).resize((224, 224))

    img_array = np.array(img)

    img_array = preprocess_input(img_array)

    img_array = np.expand_dims(img_array, axis=0)

    predictions = model.predict(img_array)

    decoded = decode_predictions(predictions, top=1)[0][0]

    breed = str(decoded[1])

    confidence = int(round(decoded[2] * 100))

    return breed, confidence
# ======================================================
# SERVE PHOTOS
# ======================================================

@app.route('/photos/<filename>')
def serve_photos(filename):
    return send_file(os.path.join('photos', filename))

# ======================================================
# HOME
# ======================================================

@app.route('/')
def home():
    return redirect(url_for('login'))

# ======================================================
# REGISTER
# ======================================================

@app.route('/register', methods=['GET', 'POST'])
def register():

    if request.method == 'POST':

        username = request.form.get('username', '').strip()
        email = request.form.get('email', '').strip()
        password = request.form.get('password', '')

        if not username:
            return render_template('register.html', message="Username is required!")
        if not email:
            return render_template('register.html', message="Email is required!")
        if not password:
            return render_template('register.html', message="Password is required!")

        success = register_user(username, email, password)

        if not success:
            return render_template('register.html',
                                   message="Email already exists!")

        return render_template('login.html',
                               message="Registration Successful! Please Login")

    return render_template('register.html')

# ======================================================
# LOGIN
# ======================================================

@app.route('/login', methods=['GET', 'POST'])
def login():

    if request.method == 'POST':

        email = request.form.get('email', '').strip()
        password = request.form.get('password', '')

        if not email:
            return render_template('login.html',
                                   message="Email is required!")
        if not password:
            return render_template('login.html',
                                   message="Password is required!")

        # Admin Login
        if email == ADMIN_EMAIL and password == ADMIN_PASSWORD:
            session['admin'] = True
            return redirect(url_for('admin_dashboard'))

        # User Login
        user = check_user(email, password)

        if user:
            session['user'] = user
            return redirect(url_for('index'))

        return render_template('login.html',
                               message="Invalid Credentials")

    return render_template('login.html')

# ======================================================
# LOGOUT
# ======================================================

@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('login'))

# ======================================================
# MAIN PAGE
# ======================================================

@app.route('/index', methods=['GET', 'POST'])
def index():

    if 'user' not in session:
        return redirect(url_for('login'))

    if request.method == 'POST':

        if 'image' not in request.files:
            return render_template('index.html',
                                   user=session['user'],
                                   message="Upload image")

        file = request.files['image']

        if file.filename == '':
            return render_template('index.html',
                                   user=session['user'],
                                   message="Select image")

        filepath = os.path.join(app.config['UPLOAD_FOLDER'],
                                file.filename)

        file.save(filepath)

        breed, confidence = predict_breed(filepath)

        size = str(get_size_category(breed))
        exact_size = get_exact_size(breed, size)

        diet = generate_diet_plan(breed, size)

        youtube_link = f"https://www.youtube.com/results?search_query={breed}+dog+training"

        amazon_link = "https://www.amazon.in/s?k=" + diet['food'].replace(" ", "+")

        session['breed'] = str(breed)
        session['confidence'] = float(confidence)
        session['size'] = str(size)
        session['exact_size'] = exact_size
        session['diet'] = dict(diet)

        return render_template(
            'result.html',
            image=filepath,
            breed=breed,
            confidence=confidence,
            diet=diet,
            size=size,
            exact_size=exact_size,
            youtube_link=youtube_link,
            amazon_link=amazon_link,
            user=session['user']
        )

    return render_template('index.html', user=session['user'])

# ======================================================
# DOWNLOAD REPORT
# ======================================================

@app.route('/download_report')
def download_report():

    if 'user' not in session:
        return redirect(url_for('login'))

    report = f"""
DOG REPORT
====================

Breed: {session['breed']}
Confidence: {session['confidence']} %

Size: {session['size']}

Food: {session['diet']['food']}
Meals: {session['diet']['meals']}
Extras: {session['diet']['extras']}

Generated:
{datetime.datetime.now()}
"""

    file_path = "dog_report.txt"

    with open(file_path, "w") as file:
        file.write(report)

    return send_file(file_path, as_attachment=True)

# ======================================================
# ADMIN LOGIN
# ======================================================

@app.route('/admin_login', methods=['GET', 'POST'])
def admin_login():

    if request.method == 'POST':

        email = request.form.get('email')
        password = request.form.get('password')

        if email == ADMIN_EMAIL and password == ADMIN_PASSWORD:
            session['admin'] = True
            return redirect(url_for('admin_dashboard'))

        return render_template("admin_login.html",
                               message="Invalid Admin Credentials")

    return render_template("admin_login.html")

# ======================================================
# ADMIN DASHBOARD
# ======================================================

@app.route('/admin')
def admin_dashboard():

    if 'admin' not in session:
        return redirect(url_for('admin_login'))

    wb = openpyxl.load_workbook(EXCEL_FILE)
    sheet = wb.active

    users = []

    for i, row in enumerate(sheet.iter_rows(min_row=2, values_only=True)):
        users.append({
            "id": i,
            "username": row[0],
            "email": row[1],
            "password": row[2]
        })

    return render_template("admin.html", users=users)

# ======================================================
# DELETE USER
# ======================================================

@app.route('/delete_user/<int:user_id>')
def delete_user(user_id):

    wb = openpyxl.load_workbook(EXCEL_FILE)
    sheet = wb.active

    sheet.delete_rows(user_id + 2)

    wb.save(EXCEL_FILE)

    return redirect(url_for('admin_dashboard'))

# ======================================================
# EDIT USER
# ======================================================

@app.route('/edit_user/<int:user_id>', methods=['GET', 'POST'])
def edit_user(user_id):

    wb = openpyxl.load_workbook(EXCEL_FILE)
    sheet = wb.active

    row_no = user_id + 2

    if request.method == 'POST':

        sheet.cell(row=row_no, column=1).value = request.form['username']
        sheet.cell(row=row_no, column=2).value = request.form['email']
        sheet.cell(row=row_no, column=3).value = request.form['password']

        wb.save(EXCEL_FILE)

        return redirect(url_for('admin_dashboard'))

    user = {
        "username": sheet.cell(row=row_no, column=1).value,
        "email": sheet.cell(row=row_no, column=2).value,
        "password": sheet.cell(row=row_no, column=3).value
    }

    return render_template("edit_user.html", user=user)

# ======================================================
# ADMIN LOGOUT
# ======================================================

@app.route('/admin_logout')
def admin_logout():
    session.pop('admin', None)
    return redirect(url_for('login'))

# ======================================================
# VACCINATIONS & GROOMING DATABASES
# ======================================================

import sqlite3

VACCINATIONS_FILE = 'vaccinations.xlsx'
GROOMING_FILE = 'grooming.xlsx'
DATABASE_FILE = 'dog_planner.db'

def get_db_connection():
    conn = sqlite3.connect(DATABASE_FILE)
    conn.row_factory = sqlite3.Row
    return conn

def init_vaccinations_db():
    if not os.path.exists(VACCINATIONS_FILE):
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.title = "Vaccinations"
        sheet.append(["Username", "PetName", "VaccineName", "Date", "Status"])
        wb.save(VACCINATIONS_FILE)
        
    conn = get_db_connection()
    conn.execute('''
        CREATE TABLE IF NOT EXISTS vaccinations (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            username TEXT NOT NULL,
            pet_name TEXT NOT NULL,
            vaccine_name TEXT NOT NULL,
            vaccination_date TEXT NOT NULL,
            status TEXT NOT NULL,
            proof_image TEXT,
            admin_remark TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    ''')
    conn.commit()
    conn.close()

def init_grooming_db():
    if not os.path.exists(GROOMING_FILE):
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.title = "Grooming"
        sheet.append(["Username", "PetName", "ServiceType", "Date", "Time", "Status"])
        wb.save(GROOMING_FILE)
        
    conn = get_db_connection()
    conn.execute('''
        CREATE TABLE IF NOT EXISTS grooming (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            username TEXT NOT NULL,
            pet_name TEXT NOT NULL,
            service_type TEXT NOT NULL,
            grooming_date TEXT NOT NULL,
            grooming_time TEXT NOT NULL,
            special_instruction TEXT,
            status TEXT NOT NULL,
            admin_remark TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    ''')
    conn.commit()
    conn.close()

def get_vaccinations(username):
    init_vaccinations_db()
    try:
        conn = get_db_connection()
        rows = conn.execute(
            'SELECT * FROM vaccinations WHERE username = ? ORDER BY vaccination_date ASC', 
            (username,)
        ).fetchall()
        conn.close()
        
        records = []
        for r in rows:
            records.append({
                "id": r["id"],
                "username": r["username"],
                "pet_name": r["pet_name"],
                "vaccine_name": r["vaccine_name"],
                "vaccination_date": r["vaccination_date"],
                "date": r["vaccination_date"],  # For backward-compatibility with health_dashboard.html
                "status": r["status"],
                "proof_image": r["proof_image"],
                "admin_remark": r["admin_remark"],
                "created_at": r["created_at"]
            })
        return records
    except Exception as e:
        print(f"Error in get_vaccinations: {e}")
        return []

def add_vaccination(username, pet_name, vaccine_name, date, status="Scheduled", proof_image=None, admin_remark=None):
    init_vaccinations_db()
    try:
        conn = get_db_connection()
        conn.execute(
            'INSERT INTO vaccinations (username, pet_name, vaccine_name, vaccination_date, status, proof_image, admin_remark) VALUES (?, ?, ?, ?, ?, ?, ?)',
            (username, pet_name, vaccine_name, date, status, proof_image, admin_remark)
        )
        conn.commit()
        conn.close()
        return True
    except Exception as e:
        print(f"Error in add_vaccination: {e}")
        return False

def get_grooming_sessions(username):
    init_grooming_db()
    try:
        conn = get_db_connection()
        rows = conn.execute(
            'SELECT * FROM grooming WHERE username = ? ORDER BY grooming_date ASC', 
            (username,)
        ).fetchall()
        conn.close()
        
        records = []
        for r in rows:
            records.append({
                "id": r["id"],
                "username": r["username"],
                "pet_name": r["pet_name"],
                "service_type": r["service_type"],
                "grooming_date": r["grooming_date"],
                "date": r["grooming_date"],  # For backward-compatibility with health_dashboard.html
                "grooming_time": r["grooming_time"],
                "grooming_time_12hr": convert_to_12_hour_format(r["grooming_time"]),  # 12-hour format
                "time": convert_to_12_hour_format(r["grooming_time"]),  # For backward-compatibility with health_dashboard.html
                "special_instruction": r["special_instruction"],
                "status": r["status"],
                "admin_remark": r["admin_remark"],
                "created_at": r["created_at"]
            })
        return records
    except Exception as e:
        print(f"Error in get_grooming_sessions: {e}")
        return []

def add_grooming_session(username, pet_name, service_type, date, time, status="Pending Approval", special_instruction="", admin_remark=None):
    init_grooming_db()
    try:
        conn = get_db_connection()
        conn.execute(
            'INSERT INTO grooming (username, pet_name, service_type, grooming_date, grooming_time, special_instruction, status, admin_remark) VALUES (?, ?, ?, ?, ?, ?, ?, ?)',
            (username, pet_name, service_type, date, time, special_instruction, status, admin_remark)
        )
        conn.commit()
        conn.close()
        return True
    except Exception as e:
        print(f"Error in add_grooming_session: {e}")
        return False

# ======================================================
# NEW ROUTES
# ======================================================

@app.route('/vaccination', methods=['GET', 'POST'])
def vaccination():
    if 'user' not in session:
        return redirect(url_for('login'))
    
    username = session['user']
    
    if request.method == 'POST':
        pet_name = request.form.get('pet_name', '').strip()
        vaccine_name = request.form.get('vaccine_name', '').strip()
        date = request.form.get('date', '').strip()
        
        # Validate required fields first
        if not pet_name or not vaccine_name or not date:
            flash("All fields are required to schedule a vaccine!", "danger")
            return redirect(url_for('vaccination'))
        
        try:
            selected_date = datetime.datetime.strptime(date, '%Y-%m-%d').date()
            
            # Check if selected date is in the past
            if selected_date < datetime.date.today():
                flash("Older dates are not allowed!", "danger")
                return redirect(url_for('vaccination'))
                
        except ValueError:
            flash("Invalid date format. Please select a valid date.", "danger")
            return redirect(url_for('vaccination'))
        
        # If validation passes, add vaccination
        success = add_vaccination(username, pet_name, vaccine_name, date)
        if success:
            flash(f"Successfully scheduled {vaccine_name} for {pet_name} on {date}!", "success")
        else:
            flash("Failed to schedule vaccination. Please try again.", "danger")
        return redirect(url_for('vaccination'))
            
    records = get_vaccinations(username)
    
    # Calculate reminders / notifications
    reminders = []
    today = datetime.date.today()
    for r in records:
        if r['status'] == 'Scheduled':
            try:
                v_date = datetime.datetime.strptime(r['vaccination_date'], '%Y-%m-%d').date()
                days_diff = (v_date - today).days
                if days_diff < 0:
                    reminders.append({
                        "type": "overdue",
                        "message": f"🚨 Overdue! {r['pet_name']}'s {r['vaccine_name']} was scheduled for {r['vaccination_date']}."
                    })
                elif 0 <= days_diff <= 7:
                    reminders.append({
                        "type": "upcoming",
                        "message": f"⏰ Upcoming! {r['pet_name']}'s {r['vaccine_name']} is due in {days_diff} days on {r['vaccination_date']}."
                    })
            except Exception as e:
                print(f"Error parsing date: {e}")
                
    return render_template(
    'vaccination.html',
    user=username,
    records=records,
    reminders=reminders,
    today=datetime.date.today().strftime('%Y-%m-%d')
)

@app.route('/vaccination/take/<int:record_id>', methods=['POST'])
def take_vaccination(record_id):
    if 'user' not in session:
        return redirect(url_for('login'))
        
    username = session['user']
    file = request.files.get('proof_image')
    
    if not file or file.filename == '':
        flash("Upload proof image/certificate to mark as taken!", "danger")
        return redirect(url_for('vaccination'))
        
    filename = f"proof_{record_id}_{file.filename}"
    filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
    file.save(filepath)
    
    conn = get_db_connection()
    record = conn.execute('SELECT * FROM vaccinations WHERE id = ? AND username = ?', (record_id, username)).fetchone()
    
    if record:
        conn.execute(
            'UPDATE vaccinations SET status = ?, proof_image = ? WHERE id = ?',
            ('Pending Approval', filename, record_id)
        )
        conn.commit()
        flash("Proof uploaded successfully! Request submitted for admin approval.", "success")
    else:
        flash("Record not found or unauthorized.", "danger")
        
    conn.close()
    return redirect(url_for('vaccination'))

@app.route('/admin/vaccinations', methods=['GET', 'POST'])
def admin_vaccinations():
    if 'admin' not in session:
        return redirect(url_for('admin_login'))
        
    conn = get_db_connection()
    
    if request.method == 'POST':
        action = request.form.get('action')
        record_id = request.form.get('record_id')
        remark = request.form.get('admin_remark', '').strip()
        
        if action == 'approve':
            conn.execute(
                'UPDATE vaccinations SET status = ?, admin_remark = ? WHERE id = ?',
                ('Approved', remark, record_id)
            )
            conn.commit()
            flash(f"Vaccination ID {record_id} successfully approved!", "success")
        elif action == 'reject':
            conn.execute(
                'UPDATE vaccinations SET status = ?, admin_remark = ? WHERE id = ?',
                ('Rejected', remark, record_id)
            )
            conn.commit()
            flash(f"Vaccination ID {record_id} rejected.", "warning")
            
    records = conn.execute('SELECT * FROM vaccinations ORDER BY created_at DESC').fetchall()
    
    # Calculate stats
    total = len(records)
    pending = conn.execute("SELECT COUNT(*) FROM vaccinations WHERE status = 'Pending Approval'").fetchone()[0]
    approved = conn.execute("SELECT COUNT(*) FROM vaccinations WHERE status = 'Approved'").fetchone()[0]
    rejected = conn.execute("SELECT COUNT(*) FROM vaccinations WHERE status = 'Rejected'").fetchone()[0]
    
    conn.close()
    
    stats = {
        "total": total,
        "pending": pending,
        "approved": approved,
        "rejected": rejected
    }
    
    return render_template('admin_vaccinations.html', records=records, stats=stats)


@app.route('/grooming', methods=['GET', 'POST'])
def grooming():
    if 'user' not in session:
        return redirect(url_for('login'))
        
    username = session['user']
    
    if request.method == 'POST':
        pet_name = request.form.get('pet_name', '').strip()
        service_type = request.form.get('service_type', '').strip()
        date = request.form.get('date', '').strip()
        time = request.form.get('time', '').strip()
        special_instruction = request.form.get('special_instruction', '').strip()
        
        # Validate required fields first
        if not pet_name or not service_type or not date or not time:
            flash("Pet Name, Service Type, Date, and Time are required!", "danger")
            return redirect(url_for('grooming'))
        
        try:
            selected_date = datetime.datetime.strptime(date, '%Y-%m-%d').date()
            
            # Check if selected date is in the past
            if selected_date < datetime.date.today():
                flash("Older dates are not allowed!", "danger")
                return redirect(url_for('grooming'))
                
        except ValueError:
            flash("Invalid date format. Please select a valid date.", "danger")
            return redirect(url_for('grooming'))
        
        # If validation passes, add grooming session
        success = add_grooming_session(
            username=username,
            pet_name=pet_name,
            service_type=service_type,
            date=date,
            time=time,
            status="Pending Approval",
            special_instruction=special_instruction
        )
        if success:
            flash(f"Successfully booked {service_type} for {pet_name}! Awaiting admin approval.", "success")
        else:
            flash("Failed to book grooming session. Please try again.", "danger")
        return redirect(url_for('grooming'))
            
    records = get_grooming_sessions(username)
    
    # Generate grooming alerts/notifications for today/tomorrow
    reminders = []
    today = datetime.date.today()
    tomorrow = today + datetime.timedelta(days=1)
    for r in records:
        if r['status'] == 'Approved':
            try:
                g_date = datetime.datetime.strptime(r['grooming_date'], '%Y-%m-%d').date()
                if g_date == today:
                    reminders.append(f"⏰ Today! {r['pet_name']} has a {r['service_type']} session at {r['grooming_time']} today.")
                elif g_date == tomorrow:
                    reminders.append(f"📅 Tomorrow! {r['pet_name']} has a {r['service_type']} session at {r['grooming_time']} tomorrow.")
            except Exception as e:
                print(f"Error parsing grooming date: {e}")
                
    return render_template('grooming.html', user=username, records=records, reminders=reminders, today=datetime.date.today().strftime('%Y-%m-%d'))

@app.route('/grooming/cancel/<int:booking_id>', methods=['POST'])
def cancel_grooming(booking_id):
    if 'user' not in session:
        return redirect(url_for('login'))
        
    username = session['user']
    
    conn = get_db_connection()
    booking = conn.execute('SELECT * FROM grooming WHERE id = ? AND username = ?', (booking_id, username)).fetchone()
    
    if booking:
        conn.execute(
            "UPDATE grooming SET status = 'Cancelled' WHERE id = ?",
            (booking_id,)
        )
        conn.commit()
        flash("Grooming appointment successfully cancelled.", "warning")
    else:
        flash("Appointment not found or unauthorized.", "danger")
        
    conn.close()
    return redirect(url_for('grooming'))

@app.route('/grooming/reschedule/<int:booking_id>', methods=['POST'])
def reschedule_grooming(booking_id):
    if 'user' not in session:
        return redirect(url_for('login'))
        
    username = session['user']
    new_date = request.form.get('date', '').strip()
    new_time = request.form.get('time', '').strip()
    
    # Validate required fields
    if not new_date or not new_time:
        flash("Date and Time are required to reschedule!", "danger")
        return redirect(url_for('grooming'))
    
    try:
        selected_date = datetime.datetime.strptime(new_date, '%Y-%m-%d').date()
        
        # Check if selected date is in the past
        if selected_date < datetime.date.today():
            flash("Older dates are not allowed!", "danger")
            return redirect(url_for('grooming'))
            
    except ValueError:
        flash("Invalid date format. Please select a valid date.", "danger")
        return redirect(url_for('grooming'))
        
    conn = get_db_connection()
    booking = conn.execute('SELECT * FROM grooming WHERE id = ? AND username = ?', (booking_id, username)).fetchone()
    
    if booking:
        conn.execute(
            "UPDATE grooming SET grooming_date = ?, grooming_time = ?, status = 'Pending Approval' WHERE id = ?",
            (new_date, new_time, booking_id)
        )
        conn.commit()
        flash("Grooming appointment successfully rescheduled! Awaiting admin approval.", "success")
    else:
        flash("Appointment not found or unauthorized.", "danger")
        
    conn.close()
    return redirect(url_for('grooming'))

@app.route('/admin/grooming', methods=['GET', 'POST'])
def admin_grooming():
    if 'admin' not in session:
        return redirect(url_for('admin_login'))
        
    conn = get_db_connection()
    
    if request.method == 'POST':
        action = request.form.get('action')
        booking_id = request.form.get('booking_id')
        remark = request.form.get('admin_remark', '').strip()
        
        if action == 'approve':
            conn.execute(
                "UPDATE grooming SET status = 'Approved', admin_remark = ? WHERE id = ?",
                (remark, booking_id)
            )
            conn.commit()
            flash(f"Grooming Session ID {booking_id} successfully approved!", "success")
        elif action == 'reject':
            conn.execute(
                "UPDATE grooming SET status = 'Rejected', admin_remark = ? WHERE id = ?",
                (remark, booking_id)
            )
            conn.commit()
            flash(f"Grooming Session ID {booking_id} rejected.", "warning")
        elif action == 'complete':
            conn.execute(
                "UPDATE grooming SET status = 'Completed', admin_remark = ? WHERE id = ?",
                (remark, booking_id)
            )
            conn.commit()
            flash(f"Grooming Session ID {booking_id} marked as completed!", "success")
            
    records_raw = conn.execute('SELECT * FROM grooming ORDER BY created_at DESC').fetchall()
    
    # Convert records to include 12-hour time format
    records = []
    for r in records_raw:
        record_dict = dict(r)
        record_dict['grooming_time_12hr'] = convert_to_12_hour_format(record_dict['grooming_time'])
        records.append(record_dict)
    
    # Calculate statistics
    total = len(records)
    pending = conn.execute("SELECT COUNT(*) FROM grooming WHERE status = 'Pending Approval'").fetchone()[0]
    approved = conn.execute("SELECT COUNT(*) FROM grooming WHERE status = 'Approved'").fetchone()[0]
    completed = conn.execute("SELECT COUNT(*) FROM grooming WHERE status = 'Completed'").fetchone()[0]
    cancelled = conn.execute("SELECT COUNT(*) FROM grooming WHERE status = 'Cancelled'").fetchone()[0]
    
    conn.close()
    
    stats = {
        "total": total,
        "pending": pending,
        "approved": approved,
        "completed": completed,
        "cancelled": cancelled
    }
    
    return render_template('admin_grooming.html', records=records, stats=stats)

@app.route('/health_dashboard')
def health_dashboard():
    if 'user' not in session:
        return redirect(url_for('login'))
        
    username = session['user']
    
    vaccinations = get_vaccinations(username)
    grooming_sessions = get_grooming_sessions(username)
    
    latest_vaccine = None
    if vaccinations:
        latest_vaccine = vaccinations[-1]
        
    next_grooming = None
    if grooming_sessions:
        next_grooming = grooming_sessions[-1]
        
    return render_template('health_dashboard.html', 
                           user=username, 
                           latest_vaccine=latest_vaccine, 
                           next_grooming=next_grooming)

@app.route('/disease_prediction', methods=['GET', 'POST'])
def disease_prediction():
    if 'user' not in session:
        return redirect(url_for('login'))
    return render_template('disease_prediction.html', user=session['user'])

# ======================================================
# EXTRA PAGES
# ======================================================

@app.route('/about')
def about():
    if 'user' not in session:
        return redirect(url_for('login'))
    return render_template('about.html', user=session['user'])

@app.route('/services')
def services():
    if 'user' not in session:
        return redirect(url_for('login'))
    return render_template('services.html', user=session['user'])

@app.route('/vets')
def vets():
    if 'user' not in session:
        return redirect(url_for('login'))
    return render_template('vets.html', user=session['user'])

def get_user_email(username):
    try:
        wb = openpyxl.load_workbook(EXCEL_FILE)
        sheet = wb.active
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if not row or len(row) < 2:
                continue
            db_user = str(row[0]).strip()
            db_email = str(row[1]).strip()
            if db_user.lower() == username.lower():
                return db_email
        return None
    except Exception as e:
        print(f"Error in get_user_email: {e}")
        return None

@app.route('/profile')
def profile():
    if 'user' not in session:
        return redirect(url_for('login'))
    username = session['user']
    email = get_user_email(username) or "No Email Found"
    return render_template('profile.html', user=username, email=email)

@app.route('/contact', methods=['GET', 'POST'])
def contact():
    if 'user' not in session:
        return redirect(url_for('login'))
    if request.method == 'POST':
        return render_template('contact.html', user=session['user'], message="Message sent successfully!")
    return render_template('contact.html', user=session['user'])

# ======================================================

if __name__ == '__main__':
    app.run(debug=True)
